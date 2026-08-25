from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from backend.app.db.database import get_db
from backend.app.models.vehicle import Vehicle


router = APIRouter(
    prefix="/vehicles",
    tags=["Vehicle Database"],
)


COLUMNS = [
    "plate_number",
    "registration_number",
    "chassis_number",
    "owner_nin_number",
    "phone_number",
    "owner_department",
    "owner_address",
    "vehicle_type",
    "make",
    "model",
    "color",
]

REQUIRED_COLUMNS = {
    "plate_number",
    "registration_number",
    "chassis_number",
}

MAX_LENGTHS = {
    "plate_number": 20,
    "registration_number": 50,
    "chassis_number": 17,
    "owner_nin_number": 20,
    "phone_number": 20,
    "owner_department": 100,
    "owner_address": 255,
    "vehicle_type": 50,
    "make": 100,
    "model": 100,
    "color": 50,
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize(value):
    return clean(value).upper()


def build_template():
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Vehicles"

    instructions = workbook.create_sheet("Instructions")

    # Header row
    for column_number, column_name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(
            row=1,
            column=column_number,
            value=column_name,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="17202A",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:L1000"
    sheet.row_dimensions[1].height = 28

    widths = {
        "A": 20,
        "B": 25,
        "C": 24,
        "D": 22,
        "E": 20,
        "F": 24,
        "G": 40,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 18,
        "L": 15,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    # Highlight required columns.
    for row in range(2, 1002):
        for column_number, column_name in enumerate(
            COLUMNS,
            start=1,
        ):
            cell = sheet.cell(
                row=row,
                column=column_number,
            )

            if column_name in REQUIRED_COLUMNS:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFF2CC",
                )

    # Vehicle type dropdown.
    validation = DataValidation(
        type="list",
        formula1='"CAR,SUV,TRUCK,VAN,BUS,MOTORCYCLE,OTHER"',
        allow_blank=True,
    )

    validation.errorTitle = "Invalid vehicle type"
    validation.error = "Select a vehicle type from the list."

    sheet.add_data_validation(validation)
    validation.add("H2:H1001")

    # Instructions sheet.
    instructions["A1"] = "National Intelligent Platform"
    instructions["A1"].font = Font(
        bold=True,
        size=16,
    )

    instructions["A3"] = "Vehicle Database Import Instructions"
    instructions["A3"].font = Font(
        bold=True,
        size=13,
    )

    lines = [
        "1. Enter one vehicle per row on the Vehicles sheet.",
        "2. Do not rename or reorder the columns.",
        "3. plate_number is required.",
        "4. registration_number is required.",
        "5. chassis_number is required.",
        "6. Do not duplicate plate numbers in the workbook.",
        "7. Do not duplicate registration numbers in the workbook.",
        "8. Do not duplicate chassis numbers in the workbook.",
        "9. Phone, department and address may be left blank when unavailable.",
        "10. Save the completed workbook as .xlsx.",
        "11. Upload the completed workbook through the website.",
        "12. The system validates the complete workbook before importing.",
        "13. If validation fails, no rows from that upload are imported.",
    ]

    for row_number, line in enumerate(lines, start=5):
        instructions.cell(
            row=row_number,
            column=1,
            value=line,
        )

    instructions.column_dimensions["A"].width = 110

    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


@router.get("/template")
def download_vehicle_template():
    output = build_template()

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; "
                'filename="national_vehicle_database_template.xlsx"'
            )
        },
    )


@router.post("/import")
async def import_vehicle_database(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = (file.filename or "").lower()

    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx Excel files are accepted.",
        )

    contents = await file.read()

    if not contents:
        raise HTTPException(
            status_code=400,
            detail="The uploaded Excel file is empty.",
        )

    try:
        workbook = load_workbook(
            BytesIO(contents),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="The uploaded file is not a valid .xlsx workbook.",
        ) from exc

    if "Vehicles" not in workbook.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="The workbook must contain a sheet named 'Vehicles'.",
        )

    sheet = workbook["Vehicles"]
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration as exc:
        raise HTTPException(
            status_code=400,
            detail="The Vehicles sheet is empty.",
        ) from exc

    received_headers = [clean(x).lower() for x in header]

    if received_headers != COLUMNS:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "The workbook columns do not match the official template.",
                "expected_columns": COLUMNS,
                "received_columns": received_headers,
            },
        )

    records = []
    errors = []

    for row_number, row in enumerate(rows, start=2):
        values = list(row)

        if not any(clean(value) for value in values):
            continue

        record = {}

        for index, column in enumerate(COLUMNS):
            value = values[index] if index < len(values) else None
            record[column] = clean(value)

        for column in REQUIRED_COLUMNS:
            if not record[column]:
                errors.append(
                    f"Row {row_number}: {column} is required."
                )

        for column, maximum in MAX_LENGTHS.items():
            if len(record[column]) > maximum:
                errors.append(
                    f"Row {row_number}: {column} exceeds {maximum} characters."
                )

        records.append(
            {
                "row_number": row_number,
                "data": record,
            }
        )

    if not records:
        raise HTTPException(
            status_code=400,
            detail="No vehicle records were found in the workbook.",
        )

    # Check duplicates inside the uploaded workbook.
    seen_plates = {}
    seen_registrations = {}
    seen_chassis = {}

    for item in records:
        row_number = item["row_number"]
        data = item["data"]

        plate = normalize(data["plate_number"])
        registration = normalize(data["registration_number"])
        chassis = normalize(data["chassis_number"])

        if plate in seen_plates:
            errors.append(
                f"Row {row_number}: duplicate plate '{plate}', "
                f"already used on row {seen_plates[plate]}."
            )
        else:
            seen_plates[plate] = row_number

        if registration in seen_registrations:
            errors.append(
                f"Row {row_number}: duplicate registration "
                f"'{registration}', already used on row "
                f"{seen_registrations[registration]}."
            )
        else:
            seen_registrations[registration] = row_number

        if chassis in seen_chassis:
            errors.append(
                f"Row {row_number}: duplicate chassis "
                f"'{chassis}', already used on row "
                f"{seen_chassis[chassis]}."
            )
        else:
            seen_chassis[chassis] = row_number

    if errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Excel validation failed. No vehicles were imported.",
                "error_count": len(errors),
                "errors": errors[:200],
            },
        )

    # Check the database before importing.
    plate_values = list(seen_plates.keys())
    registration_values = list(seen_registrations.keys())
    chassis_values = list(seen_chassis.keys())

    existing = db.scalars(
        select(Vehicle).where(
            or_(
                Vehicle.plate_number.in_(plate_values),
                Vehicle.registration_number.in_(registration_values),
                Vehicle.chassis_number.in_(chassis_values),
            )
        )
    ).all()

    existing_plates = {
        normalize(vehicle.plate_number)
        for vehicle in existing
    }

    existing_registrations = {
        normalize(vehicle.registration_number)
        for vehicle in existing
    }

    existing_chassis = {
        normalize(vehicle.chassis_number)
        for vehicle in existing
    }

    database_errors = []

    for item in records:
        row_number = item["row_number"]
        data = item["data"]

        plate = normalize(data["plate_number"])
        registration = normalize(data["registration_number"])
        chassis = normalize(data["chassis_number"])

        if plate in existing_plates:
            database_errors.append(
                f"Row {row_number}: plate '{plate}' already exists."
            )

        if registration in existing_registrations:
            database_errors.append(
                f"Row {row_number}: registration "
                f"'{registration}' already exists."
            )

        if chassis in existing_chassis:
            database_errors.append(
                f"Row {row_number}: chassis '{chassis}' already exists."
            )

    if database_errors:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Database duplicate check failed. No vehicles were imported.",
                "error_count": len(database_errors),
                "errors": database_errors[:200],
            },
        )

    # Create all objects first, then commit once.
    vehicles = []

    for item in records:
        data = item["data"]

        vehicles.append(
            Vehicle(
                plate_number=normalize(data["plate_number"]),
                registration_number=normalize(data["registration_number"]),
                chassis_number=normalize(data["chassis_number"]),
                owner_nin_number=data["owner_nin_number"] or None,
                phone_number=data["phone_number"] or None,
                owner_department=data["owner_department"] or None,
                owner_address=data["owner_address"] or None,
                vehicle_type=data["vehicle_type"] or None,
                make=data["make"] or None,
                model=data["model"] or None,
                color=data["color"] or None,
            )
        )

    try:
        db.add_all(vehicles)
        db.commit()

    except IntegrityError as exc:
        db.rollback()

        raise HTTPException(
            status_code=409,
            detail=(
                "Database rejected the import because of a "
                "duplicate or constraint violation. "
                "No vehicles were imported."
            ),
        ) from exc

    except Exception as exc:
        db.rollback()

        raise HTTPException(
            status_code=500,
            detail="Vehicle import failed. No vehicles were imported.",
        ) from exc

    return {
        "status": "IMPORT_COMPLETE",
        "filename": file.filename,
        "vehicles_imported": len(vehicles),
    }
